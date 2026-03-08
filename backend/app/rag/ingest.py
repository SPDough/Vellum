"""
Otomeshon RAG Ingestion Pipeline
=================================
Processes documents into chunks, generates embeddings, and loads into pgvector.

Supported document types:
  - PDF books / textbooks     (pypdf + chapter/section-aware chunking)
  - Excel API specs            (per-mart field grouping)
  - PDF operational manuals    (DTCC/NSCC/DTC — procedure-aware chunking)
  - PDF academic papers        (abstract/section-aware chunking)

Chunking strategies by doc_type:
  'book'    → section_boundary chunking: respects CHAPTER / Roman-numeral section headers
  'api_spec'→ mart chunking:             one chunk group per data mart
  'manual'  → procedure chunking:        respects numbered sections, rule IDs, tables
  'paper'   → paper chunking:            respects Abstract/Introduction/etc. headings

Usage:
  python ingest.py --doc weiss_book
  python ingest.py --doc custody_api
  python ingest.py --doc performance_api
  python ingest.py --all
  python ingest.py --doc dtcc_settlement_guide   # example future manual
"""

import os
import re
import json
import time
import hashlib
import argparse
import logging
from dataclasses import dataclass, field, asdict
from typing import Optional
from datetime import datetime, timezone

import psycopg2
from psycopg2.extras import execute_values
import openai
import tiktoken
from pypdf import PdfReader
import openpyxl

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger(__name__)

# ── Configuration ─────────────────────────────────────────────────────────────
EMBEDDING_MODEL     = "text-embedding-3-small"   # 1536 dims, best cost/quality tradeoff
EMBEDDING_DIMS      = 1536
CHUNK_TARGET_TOKENS = 400    # target chunk size — balances context density vs retrieval precision
CHUNK_OVERLAP_TOKENS= 60     # overlap between consecutive chunks to preserve context across boundaries
BATCH_SIZE          = 50     # embeddings per API call (OpenAI max is 2048)

def _get_config():
    """DB DSN and OpenAI key from app config or env."""
    try:
        from app.core.config import get_settings
        s = get_settings()
        dsn = getattr(s, "database_url", None)
        key = getattr(s, "openai_api_key", None)
        if dsn and dsn.startswith("postgresql+psycopg"):
            dsn = dsn.replace("postgresql+psycopg", "postgresql", 1)
        return dsn or os.environ.get("OTOMESHON_DB_DSN") or os.environ.get("DATABASE_URL"), key or os.environ.get("OPENAI_API_KEY")
    except Exception:
        return os.environ.get("OTOMESHON_DB_DSN") or os.environ.get("DATABASE_URL"), os.environ.get("OPENAI_API_KEY")

DB_DSN = None  # set in main() from _get_config()
OPENAI_API_KEY = None
# Table names for NAV RAG (migration 003)
NAV_RAG_DOCS_TABLE = "nav_rag_documents"
NAV_RAG_CHUNKS_TABLE = "nav_rag_chunks"


# ── Data model ────────────────────────────────────────────────────────────────
@dataclass
class DocumentRecord:
    doc_id:           str
    title:            str
    doc_type:         str       # 'book' | 'api_spec' | 'manual' | 'paper'
    source_file:      str
    author:           Optional[str] = None
    publisher:        Optional[str] = None
    publication_year: Optional[int] = None
    domain:           str = "fund_operations"
    metadata:         dict = field(default_factory=dict)


@dataclass
class ChunkRecord:
    chunk_id:       str
    doc_id:         str
    content:        str
    content_tokens: int
    chunk_seq:      int
    chunk_type:     str         # 'prose' | 'api_field' | 'api_table' | 'definition'
    chapter:        Optional[str] = None
    chapter_title:  Optional[str] = None
    section:        Optional[str] = None
    section_title:  Optional[str] = None
    page_start:     Optional[int] = None
    page_end:       Optional[int] = None
    mart_name:      Optional[str] = None
    topics:         list = field(default_factory=list)
    relevance_tags: list = field(default_factory=list)
    embedding:      Optional[list] = None


# ── Token counting ────────────────────────────────────────────────────────────
_enc = tiktoken.get_encoding("cl100k_base")

def count_tokens(text: str) -> int:
    return len(_enc.encode(text))

def truncate_to_tokens(text: str, max_tokens: int) -> str:
    tokens = _enc.encode(text)
    if len(tokens) <= max_tokens:
        return text
    return _enc.decode(tokens[:max_tokens])


# ── Chunking utilities ────────────────────────────────────────────────────────
def split_into_sentences(text: str) -> list[str]:
    """Split text into sentences, preserving structure."""
    # Split on sentence-ending punctuation followed by whitespace + capital
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', text)
    return [s.strip() for s in sentences if s.strip()]


def chunk_text(
    text: str,
    target_tokens: int = CHUNK_TARGET_TOKENS,
    overlap_tokens: int = CHUNK_OVERLAP_TOKENS,
) -> list[str]:
    """
    Split text into overlapping chunks of approximately target_tokens.
    Respects sentence boundaries — never splits mid-sentence.
    Overlap preserves context across chunk boundaries.
    """
    sentences = split_into_sentences(text)
    chunks = []
    current_sentences = []
    current_tokens = 0

    for sentence in sentences:
        s_tokens = count_tokens(sentence)

        # If single sentence exceeds target, include it as its own chunk
        if s_tokens > target_tokens:
            if current_sentences:
                chunks.append(" ".join(current_sentences))
            chunks.append(sentence)
            current_sentences = []
            current_tokens = 0
            continue

        if current_tokens + s_tokens > target_tokens and current_sentences:
            # Flush current chunk
            chunks.append(" ".join(current_sentences))

            # Build overlap: take sentences from end of current chunk
            overlap_sentences = []
            overlap_count = 0
            for s in reversed(current_sentences):
                t = count_tokens(s)
                if overlap_count + t > overlap_tokens:
                    break
                overlap_sentences.insert(0, s)
                overlap_count += t

            current_sentences = overlap_sentences
            current_tokens = overlap_count

        current_sentences.append(sentence)
        current_tokens += s_tokens

    if current_sentences:
        chunks.append(" ".join(current_sentences))

    return [c for c in chunks if c.strip()]


# ── Chapter/section detection ─────────────────────────────────────────────────
CHAPTER_PATTERN = re.compile(
    r'CHAPTER\s+(I{1,3}V?|VI{0,3}|IX|X{1,3})\b',
    re.IGNORECASE
)
SECTION_PATTERN = re.compile(
    r'\b([IVX]+)-([A-Z](?:-\d+(?:-[a-z]+(?:-[a-z]+)?)?)?)\s+([A-Z][A-Z\s/&,]{3,60})',
)

CHAPTER_TITLES = {
    "I":    "Overview of the Industry",
    "II":   "Products That We Offer",
    "III":  "The Marketplaces and Order Management",
    "IV":   "The Trade-Processing System",
    "V":    "Margin",
    "VI":   "Settlement",
    "VII":  "Custodial Services",
    "VIII": "Role of the Banks",
    "IX":   "Cash Accounting",
    "X":    "Brokerage Accounting",
    "XI":   "Stock Record",
    "XII":  "General Ledger",
    "XIII": "Regulatory Reporting",
    "XIV":  "Conclusion",
}

def parse_chapter_section(text: str) -> tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """Extract chapter and section identifiers from text."""
    chapter = section = chapter_title = section_title = None

    m = CHAPTER_PATTERN.search(text)
    if m:
        chapter = f"CHAPTER {m.group(1).upper()}"
        chapter_title = CHAPTER_TITLES.get(m.group(1).upper())

    m2 = SECTION_PATTERN.search(text)
    if m2:
        section = f"{m2.group(1)}-{m2.group(2)}"
        section_title = m2.group(3).strip().title()

    return chapter, chapter_title, section, section_title


# ── Manual section detection (DTCC / NSCC / DTC operational documents) ────────
# Matches numbered-section formats common in clearing corp documents:
#   "3.1 Settlement Obligations"  "Section 4.2.1"  "Rule 11"  "Procedure 7-A"
MANUAL_SECTION_PATTERN = re.compile(
    r'^(?:Section\s+|Rule\s+|Procedure\s+)?(\d+(?:\.\d+){0,3})\s{1,4}([A-Z][^\n]{5,80})',
    re.MULTILINE
)
# Matches DTCC/NSCC rule citation format: "Rule 387", "Procedure 10", "Article III Section 2"
RULE_CITATION_PATTERN = re.compile(
    r'\b(?:Rule|Procedure|Article|Section)\s+\d+(?:[A-Z]|-\d+)?\b',
    re.IGNORECASE
)

def parse_manual_section(text: str) -> tuple[Optional[str], Optional[str]]:
    """
    Extract section number and title from operational manual text.
    Returns (section_id, section_title) e.g. ("3.1", "Settlement Obligations")
    """
    m = MANUAL_SECTION_PATTERN.search(text)
    if m:
        return m.group(1), m.group(2).strip()[:80]
    return None, None


# ── Academic paper section detection ─────────────────────────────────────────
# Standard IMRaD structure plus common finance paper sections
PAPER_SECTION_HEADINGS = re.compile(
    r'^(Abstract|Introduction|Literature\s+Review|Methodology|Data|'
    r'Results?|Discussion|Conclusion|References?|Appendix'
    r'|Related\s+Work|Background|Framework|Empirical\s+Analysis'
    r'|Regulatory\s+Context|Market\s+Structure|Robustness)\b',
    re.MULTILINE | re.IGNORECASE
)

def parse_paper_section(text: str) -> Optional[str]:
    """Identify standard academic paper section from text."""
    m = PAPER_SECTION_HEADINGS.search(text)
    return m.group(1).title() if m else None


# ── Topic tagging ─────────────────────────────────────────────────────────────
TOPIC_KEYWORDS = {
    "dividend":           ["dividend", "ex-dividend", "ex dividend", "payable date", "record date",
                           "declaration date", "dividend income", "DIVINC", "dividend control"],
    "corporate_action":   ["corporate action", "reorg", "reorganization", "merger", "acquisition",
                           "tender offer", "stock split", "spinoff", "spin-off", "takeoff"],
    "nav":                ["net asset value", "NAV", "nav per unit", "fund accounting", "fund NAV",
                           "navTotalAmount", "navPerUnit"],
    "settlement":         ["settlement", "settle", "T+1", "T+2", "T+3", "fail to deliver",
                           "fail to receive", "DTC", "DTCC", "NSCC", "continuous net settlement"],
    "custody":            ["custodian", "custodial", "custody", "safekeeping", "nominee",
                           "street name", "beneficial owner", "CEDE"],
    "cash_accounting":    ["cash accounting", "debit", "credit", "journal entry", "ledger",
                           "trial balance", "control account", "accounts receivable", "accounts payable"],
    "position":           ["position", "long", "short", "units", "shares", "lot",
                           "stock record", "takeoff"],
    "api_field":          ["entityId", "navTotalAmount", "navPerUnit", "transactionTypeCode",
                           "dividendAppliedIndicator", "effectiveDate", "ENTITY_NAV", "TRANSACTION",
                           "POSITION_LOT", "CASH_ACTIVITY", "CASH_BALANCE", "POSITION_SECURITY_LEVEL"],
    "paying_agent":       ["paying agent", "dividend disbursing agent", "DDA", "transfer agent",
                           "registered holder"],
    "income_accrual":     ["accrual", "accrue", "income receivable", "interest accrued",
                           "receivable", "income payable"],
    "error_detection":    ["error", "discrepancy", "missing", "not processed", "omission",
                           "reconciliation", "validation", "failed", "undetected"],
    # New topics for manuals and papers
    "regulation":         ["regulation", "regulatory", "rule", "requirement", "obligation",
                           "compliance", "Rule 387", "15c3", "Regulation T", "SEC", "FINRA"],
    "procedure":          ["procedure", "process", "workflow", "step", "instruction",
                           "notification", "submission", "filing", "report"],
    "research":           ["study", "analysis", "evidence", "hypothesis", "empirical",
                           "literature", "findings", "methodology", "data", "model"],
    "risk":               ["risk", "exposure", "counterparty", "default", "margin call",
                           "haircut", "collateral", "mark to market", "systemic"],
}

RELEVANCE_TAG_MAP = {
    "nav_validation":    ["dividend", "nav", "corporate_action", "income_accrual", "error_detection"],
    "settlement_ops":    ["settlement", "custody", "paying_agent"],
    "fund_accounting":   ["cash_accounting", "nav", "income_accrual", "position"],
    "api_reference":     ["api_field"],
    "corporate_actions": ["corporate_action", "dividend", "paying_agent"],
    # New tags for the expanded use cases
    "validation_rules":  ["regulation", "procedure", "error_detection", "settlement"],
    "client_reporting":  ["nav", "dividend", "corporate_action", "income_accrual"],
    "regulatory_basis":  ["regulation", "risk", "research"],
}

def tag_chunk(text: str) -> tuple[list[str], list[str]]:
    """Return (topics, relevance_tags) for a chunk based on keyword matching."""
    text_lower = text.lower()
    topics = [
        topic for topic, keywords in TOPIC_KEYWORDS.items()
        if any(kw.lower() in text_lower for kw in keywords)
    ]
    relevance_tags = [
        tag for tag, required_topics in RELEVANCE_TAG_MAP.items()
        if any(t in topics for t in required_topics)
    ]
    return topics, relevance_tags


# ── Embedding generation ──────────────────────────────────────────────────────
def generate_embeddings(texts: list[str], client: openai.OpenAI) -> list[list[float]]:
    """
    Generate embeddings in batches. Handles rate limits with exponential backoff.
    Returns list of embedding vectors in same order as input texts.
    """
    all_embeddings = []
    for i in range(0, len(texts), BATCH_SIZE):
        batch = texts[i:i + BATCH_SIZE]
        for attempt in range(5):
            try:
                response = client.embeddings.create(
                    model=EMBEDDING_MODEL,
                    input=batch,
                    dimensions=EMBEDDING_DIMS,
                )
                all_embeddings.extend([e.embedding for e in response.data])
                log.info(f"  Embedded batch {i//BATCH_SIZE + 1} ({len(batch)} chunks)")
                break
            except openai.RateLimitError:
                wait = 2 ** attempt
                log.warning(f"  Rate limited, waiting {wait}s...")
                time.sleep(wait)
            except Exception as e:
                log.error(f"  Embedding error: {e}")
                raise
        else:
            raise RuntimeError(f"Failed to embed batch after 5 attempts")

    return all_embeddings


# ═════════════════════════════════════════════════════════════════════════════
# PDF INGESTION — Weiss book and similar structured texts
# ═════════════════════════════════════════════════════════════════════════════

def ingest_pdf_book(
    path: str,
    doc: DocumentRecord,
    client: openai.OpenAI,
    conn,
) -> int:
    """
    Ingest a PDF book with chapter/section awareness.

    Strategy:
      1. Extract all pages, tracking chapter/section transitions
      2. Accumulate text until a natural boundary (chapter/section change or
         token threshold) then chunk within that boundary
      3. Preserve chapter + section metadata on every chunk
    """
    log.info(f"Reading PDF: {path}")
    reader = PdfReader(path)
    total_pages = len(reader.pages)
    log.info(f"  {total_pages} pages")

    # ── Pass 1: extract pages with location metadata ──────────────────────
    pages = []
    current_chapter = current_chapter_title = None
    current_section = current_section_title = None

    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if not text.strip():
            continue

        # Update chapter/section tracking
        ch, ch_title, sec, sec_title = parse_chapter_section(text)
        if ch:
            current_chapter = ch
            current_chapter_title = ch_title
        if sec:
            current_section = sec
            current_section_title = sec_title

        pages.append({
            "page_num":      i + 1,
            "text":          text,
            "chapter":       current_chapter,
            "chapter_title": current_chapter_title,
            "section":       current_section,
            "section_title": current_section_title,
        })

    # ── Pass 2: group pages into section-coherent text blocks ─────────────
    # A new block starts whenever chapter or section changes
    blocks = []
    current_block = None

    for p in pages:
        key = (p["chapter"], p["section"])
        if current_block is None or key != current_block["key"]:
            if current_block:
                blocks.append(current_block)
            current_block = {
                "key":           key,
                "chapter":       p["chapter"],
                "chapter_title": p["chapter_title"],
                "section":       p["section"],
                "section_title": p["section_title"],
                "page_start":    p["page_num"],
                "page_end":      p["page_num"],
                "text":          p["text"],
            }
        else:
            current_block["text"] += "\n" + p["text"]
            current_block["page_end"] = p["page_num"]

    if current_block:
        blocks.append(current_block)

    log.info(f"  Found {len(blocks)} section blocks")

    # ── Pass 3: chunk each block ──────────────────────────────────────────
    all_chunks: list[ChunkRecord] = []
    seq = 0

    for block in blocks:
        raw_chunks = chunk_text(block["text"])

        for chunk_text_val in raw_chunks:
            tokens = count_tokens(chunk_text_val)
            if tokens < 30:   # skip very small fragments
                continue

            topics, tags = tag_chunk(chunk_text_val)

            chunk = ChunkRecord(
                chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                doc_id         = doc.doc_id,
                content        = chunk_text_val,
                content_tokens = tokens,
                chunk_seq      = seq,
                chunk_type     = "prose",
                chapter        = block["chapter"],
                chapter_title  = block["chapter_title"],
                section        = block["section"],
                section_title  = block["section_title"],
                page_start     = block["page_start"],
                page_end       = block["page_end"],
                topics         = topics,
                relevance_tags = tags,
            )
            all_chunks.append(chunk)
            seq += 1

    log.info(f"  Generated {len(all_chunks)} chunks")

    # ── Pass 4: embed and store ───────────────────────────────────────────
    return _embed_and_store(all_chunks, doc, client, conn)


# ═════════════════════════════════════════════════════════════════════════════
# PDF INGESTION — Operational manuals (DTCC / NSCC / DTC / custodian guides)
# ═════════════════════════════════════════════════════════════════════════════

def ingest_pdf_manual(
    path: str,
    doc: DocumentRecord,
    client: openai.OpenAI,
    conn,
) -> int:
    """
    Ingest a PDF operational manual with procedure/rule-aware chunking.

    Structural differences from books:
      - Numbered sections (3.1, 4.2.1) instead of Roman numeral chapters
      - Rule citations (Rule 387, Procedure 10) should be preserved intact
      - Dense tables and step-by-step procedures — tighter chunks (300 tokens)
        so individual rules stay retrievable without excess surrounding text
      - Section number tracked as section_id, section title as section_title

    Strategy:
      1. Extract pages, detect numbered section boundaries
      2. Annotate each page with its current section
      3. Group into section blocks, chunk within each block at 300 tokens
         (tighter than books — procedure steps shouldn't be split across chunks)
    """
    MANUAL_CHUNK_TOKENS = 300   # tighter than prose — procedures are dense

    log.info(f"Reading manual PDF: {path}")
    reader   = PdfReader(path)
    total_pg = len(reader.pages)
    log.info(f"  {total_pg} pages")

    pages = []
    current_section = current_section_title = None

    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if not text.strip():
            continue

        sec_id, sec_title = parse_manual_section(text)
        if sec_id:
            current_section       = sec_id
            current_section_title = sec_title

        pages.append({
            "page_num":      i + 1,
            "text":          text,
            "section":       current_section,
            "section_title": current_section_title,
        })

    # Group by section
    blocks = []
    current_block = None

    for p in pages:
        key = p["section"]
        if current_block is None or key != current_block["key"]:
            if current_block:
                blocks.append(current_block)
            current_block = {
                "key":           key,
                "section":       p["section"],
                "section_title": p["section_title"],
                "page_start":    p["page_num"],
                "page_end":      p["page_num"],
                "text":          p["text"],
            }
        else:
            current_block["text"] += "\n" + p["text"]
            current_block["page_end"] = p["page_num"]

    if current_block:
        blocks.append(current_block)

    log.info(f"  Found {len(blocks)} section blocks")

    all_chunks: list[ChunkRecord] = []
    seq = 0

    for block in blocks:
        raw_chunks = chunk_text(
            block["text"],
            target_tokens  = MANUAL_CHUNK_TOKENS,
            overlap_tokens = 40,
        )
        for chunk_text_val in raw_chunks:
            tokens = count_tokens(chunk_text_val)
            if tokens < 20:
                continue

            # Add rule citation prefix if present — critical for retrieval
            rule_cites = RULE_CITATION_PATTERN.findall(chunk_text_val)
            topics, tags = tag_chunk(chunk_text_val)
            if rule_cites:
                topics = list(set(topics + ["regulation", "procedure"]))
                tags   = list(set(tags   + ["validation_rules", "regulatory_basis"]))

            chunk = ChunkRecord(
                chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                doc_id         = doc.doc_id,
                content        = chunk_text_val,
                content_tokens = tokens,
                chunk_seq      = seq,
                chunk_type     = "procedure",
                section        = block["section"],
                section_title  = block["section_title"],
                page_start     = block["page_start"],
                page_end       = block["page_end"],
                topics         = topics,
                relevance_tags = tags,
            )
            all_chunks.append(chunk)
            seq += 1

    log.info(f"  Generated {len(all_chunks)} chunks")
    return _embed_and_store(all_chunks, doc, client, conn)


# ═════════════════════════════════════════════════════════════════════════════
# PDF INGESTION — Academic papers (fund accounting, settlement, market structure)
# ═════════════════════════════════════════════════════════════════════════════

def ingest_pdf_paper(
    path: str,
    doc: DocumentRecord,
    client: openai.OpenAI,
    conn,
) -> int:
    """
    Ingest an academic paper with IMRaD-aware chunking.

    Key differences from books:
      - Abstract should be a standalone chunk — it summarises the whole paper
        and is often the highest-signal chunk for retrieval
      - Section headings (Introduction, Methodology, Results, etc.) mark
        thematic shifts that should start new chunks
      - References section is skipped — it's noise for retrieval
      - Figures and tables often extract as garbled text — filtered out

    Strategy:
      1. Detect Abstract and assign as chunk_type='abstract'
      2. Detect major section headings and group text accordingly
      3. Skip References section
      4. Chunk within sections at standard 400 tokens
    """
    log.info(f"Reading paper PDF: {path}")
    reader   = PdfReader(path)
    total_pg = len(reader.pages)
    log.info(f"  {total_pg} pages")

    pages = []
    current_section = "Introduction"   # default if no heading detected
    in_references   = False

    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if not text.strip():
            continue

        sec = parse_paper_section(text)
        if sec:
            # Stop ingesting at References — pure citation noise
            if sec.lower() in ("references", "reference"):
                in_references = True
            else:
                in_references   = False
                current_section = sec

        if in_references:
            continue

        pages.append({
            "page_num": i + 1,
            "text":     text,
            "section":  current_section,
        })

    # Group by section
    blocks = []
    current_block = None

    for p in pages:
        if current_block is None or p["section"] != current_block["section"]:
            if current_block:
                blocks.append(current_block)
            current_block = {
                "section":    p["section"],
                "page_start": p["page_num"],
                "page_end":   p["page_num"],
                "text":       p["text"],
            }
        else:
            current_block["text"] += "\n" + p["text"]
            current_block["page_end"] = p["page_num"]

    if current_block:
        blocks.append(current_block)

    log.info(f"  Found {len(blocks)} section blocks (references excluded)")

    all_chunks: list[ChunkRecord] = []
    seq = 0

    for block in blocks:
        is_abstract = block["section"].lower() == "abstract"

        # Abstract: single chunk regardless of length — it's the paper's summary
        if is_abstract:
            content = block["text"].strip()
            if content:
                topics, tags = tag_chunk(content)
                tags = list(set(tags + ["regulatory_basis"]))
                chunk = ChunkRecord(
                    chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                    doc_id         = doc.doc_id,
                    content        = content,
                    content_tokens = count_tokens(content),
                    chunk_seq      = seq,
                    chunk_type     = "abstract",
                    section        = "Abstract",
                    section_title  = "Abstract",
                    page_start     = block["page_start"],
                    page_end       = block["page_end"],
                    topics         = topics,
                    relevance_tags = tags,
                )
                all_chunks.append(chunk)
                seq += 1
            continue

        # All other sections: standard chunking
        raw_chunks = chunk_text(block["text"])
        for chunk_text_val in raw_chunks:
            tokens = count_tokens(chunk_text_val)
            if tokens < 30:
                continue
            topics, tags = tag_chunk(chunk_text_val)
            tags = list(set(tags + ["research"]))

            chunk = ChunkRecord(
                chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                doc_id         = doc.doc_id,
                content        = chunk_text_val,
                content_tokens = tokens,
                chunk_seq      = seq,
                chunk_type     = "prose",
                section        = block["section"],
                section_title  = block["section"],
                page_start     = block["page_start"],
                page_end       = block["page_end"],
                topics         = topics,
                relevance_tags = tags,
            )
            all_chunks.append(chunk)
            seq += 1

    log.info(f"  Generated {len(all_chunks)} chunks")
    return _embed_and_store(all_chunks, doc, client, conn)




def ingest_excel_api_spec(
    path: str,
    doc: DocumentRecord,
    client: openai.OpenAI,
    conn,
) -> int:
    """
    Ingest an Excel API spec (State Street consumption model format).

    Strategy:
      Each mart (ENTITY_NAV, TRANSACTION, etc.) becomes a set of chunks.
      Fields are grouped into chunks of ~CHUNK_TARGET_TOKENS each.
      Each chunk is a self-contained description of a group of related fields.

      Format per chunk:
        API Mart: {mart_name} ({business_service})
        Fields:
          {FIELD_NAME} ({api_attribute}): {definition} [{data_type}]
          ...
    """
    log.info(f"Reading Excel: {path}")
    wb = openpyxl.load_workbook(path, read_only=True)

    all_chunks: list[ChunkRecord] = []
    seq = 0

    for sheet_name in wb.sheetnames:
        if sheet_name not in ("Consumption Mart Dictionary", "API Listing"):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if sheet_name == "API Listing":
            # Store the API listing as a single chunk
            api_lines = [f"API Listing for {doc.title}:", ""]
            for r in rows[2:]:
                if not any(c for c in r if c):
                    continue
                vals = [str(c) for c in r if c]
                api_lines.append(" | ".join(vals))

            content = "\n".join(api_lines)
            topics, tags = tag_chunk(content)
            tags = list(set(tags + ["api_reference"]))

            chunk = ChunkRecord(
                chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                doc_id         = doc.doc_id,
                content        = content,
                content_tokens = count_tokens(content),
                chunk_seq      = seq,
                chunk_type     = "api_table",
                section        = "API_LISTING",
                section_title  = "Available APIs",
                topics         = topics,
                relevance_tags = tags,
            )
            all_chunks.append(chunk)
            seq += 1
            continue

        # Consumption Mart Dictionary
        # Group rows by mart name
        mart_rows: dict[str, list] = {}
        headers = None

        for i, row in enumerate(rows):
            if i == 0:
                continue  # title row
            if i == 1:
                headers = row
                continue
            if not any(c for c in row if c):
                continue

            business_service = str(row[0]) if row[0] else ""
            mart = str(row[1]) if row[1] else ""
            if not mart:
                continue

            key = f"{business_service}::{mart}"
            if key not in mart_rows:
                mart_rows[key] = []
            mart_rows[key].append(row)

        log.info(f"  Found {len(mart_rows)} marts in {sheet_name}")

        # Chunk each mart's fields
        for mart_key, field_rows in mart_rows.items():
            business_service, mart_name = mart_key.split("::", 1)

            # Build text for all fields in this mart
            all_field_lines = [
                f"API Mart: {mart_name}",
                f"Business Service: {business_service}",
                f"Source: {doc.title}",
                "Fields:",
            ]
            for row in field_rows:
                physical_name = str(row[2]) if row[2] else ""
                api_attr      = str(row[3]) if row[3] else physical_name
                definition    = str(row[4]) if row[4] else ""
                data_type     = str(row[5]) if row[5] else ""
                comment       = str(row[7]) if len(row) > 7 and row[7] else ""

                line = f"  {physical_name}"
                if api_attr and api_attr != physical_name:
                    line += f" (API: {api_attr})"
                if definition:
                    line += f": {definition}"
                if data_type:
                    line += f" [{data_type}]"
                if comment:
                    line += f" — Note: {comment}"
                all_field_lines.append(line)

            full_text = "\n".join(all_field_lines)
            full_tokens = count_tokens(full_text)

            if full_tokens <= CHUNK_TARGET_TOKENS * 2:
                # Small mart — store as single chunk
                topics, tags = tag_chunk(full_text)
                tags = list(set(tags + ["api_reference"]))
                chunk = ChunkRecord(
                    chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                    doc_id         = doc.doc_id,
                    content        = full_text,
                    content_tokens = full_tokens,
                    chunk_seq      = seq,
                    chunk_type     = "api_field",
                    mart_name      = mart_name,
                    section        = mart_name,
                    section_title  = f"{mart_name} Fields",
                    topics         = topics,
                    relevance_tags = tags,
                )
                all_chunks.append(chunk)
                seq += 1
            else:
                # Large mart (TRANSACTION, POSITION_LOT, etc.) — chunk by field groups
                header_lines = all_field_lines[:4]  # mart header
                field_lines  = all_field_lines[4:]

                current_lines = list(header_lines)
                current_tokens = count_tokens("\n".join(current_lines))
                chunk_num = 0

                for line in field_lines:
                    line_tokens = count_tokens(line)
                    if current_tokens + line_tokens > CHUNK_TARGET_TOKENS and len(current_lines) > 4:
                        content = "\n".join(current_lines)
                        topics, tags = tag_chunk(content)
                        tags = list(set(tags + ["api_reference"]))
                        chunk = ChunkRecord(
                            chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                            doc_id         = doc.doc_id,
                            content        = content,
                            content_tokens = count_tokens(content),
                            chunk_seq      = seq,
                            chunk_type     = "api_field",
                            mart_name      = mart_name,
                            section        = mart_name,
                            section_title  = f"{mart_name} Fields (part {chunk_num + 1})",
                            topics         = topics,
                            relevance_tags = tags,
                        )
                        all_chunks.append(chunk)
                        seq += 1
                        chunk_num += 1
                        # Reset with header + overlap
                        current_lines = list(header_lines) + current_lines[-3:]
                        current_tokens = count_tokens("\n".join(current_lines))

                    current_lines.append(line)
                    current_tokens += line_tokens

                # Flush final chunk
                if len(current_lines) > 4:
                    content = "\n".join(current_lines)
                    topics, tags = tag_chunk(content)
                    tags = list(set(tags + ["api_reference"]))
                    chunk = ChunkRecord(
                        chunk_id       = f"{doc.doc_id}::chunk::{seq:04d}",
                        doc_id         = doc.doc_id,
                        content        = content,
                        content_tokens = count_tokens(content),
                        chunk_seq      = seq,
                        chunk_type     = "api_field",
                        mart_name      = mart_name,
                        section        = mart_name,
                        section_title  = f"{mart_name} Fields (part {chunk_num + 1})",
                        topics         = topics,
                        relevance_tags = tags,
                    )
                    all_chunks.append(chunk)
                    seq += 1

    log.info(f"  Generated {len(all_chunks)} chunks")
    return _embed_and_store(all_chunks, doc, client, conn)


# ═════════════════════════════════════════════════════════════════════════════
# DATABASE WRITE
# ═════════════════════════════════════════════════════════════════════════════

def _embed_and_store(
    chunks: list[ChunkRecord],
    doc: DocumentRecord,
    client: openai.OpenAI,
    conn,
) -> int:
    """Embed all chunks and write document + chunks to Postgres."""

    # Generate embeddings
    log.info(f"  Generating embeddings for {len(chunks)} chunks...")
    texts = [c.content for c in chunks]
    embeddings = generate_embeddings(texts, client)

    for chunk, emb in zip(chunks, embeddings):
        chunk.embedding = emb

    # Write to database (NAV RAG tables)
    docs_t = NAV_RAG_DOCS_TABLE
    chunks_t = NAV_RAG_CHUNKS_TABLE
    with conn.cursor() as cur:
        # Upsert document record
        cur.execute(f"""
            INSERT INTO {docs_t}
                (doc_id, title, doc_type, source_file, author, publisher,
                 publication_year, domain, total_chunks, total_tokens, metadata)
            VALUES
                (%(doc_id)s, %(title)s, %(doc_type)s, %(source_file)s, %(author)s,
                 %(publisher)s, %(publication_year)s, %(domain)s, %(total_chunks)s,
                 %(total_tokens)s, %(metadata)s)
            ON CONFLICT (doc_id) DO UPDATE SET
                total_chunks = EXCLUDED.total_chunks,
                total_tokens = EXCLUDED.total_tokens,
                ingested_at  = NOW()
        """, {
            **asdict(doc),
            "total_chunks": len(chunks),
            "total_tokens": sum(c.content_tokens for c in chunks),
            "metadata":     json.dumps(doc.metadata),
        })

        # Delete existing chunks for this doc (clean re-ingest)
        cur.execute(f"DELETE FROM {chunks_t} WHERE doc_id = %s", (doc.doc_id,))

        # Bulk insert chunks
        chunk_rows = [
            (
                c.chunk_id, c.doc_id, c.content, c.content_tokens,
                c.chunk_seq, c.chunk_type,
                c.chapter, c.chapter_title,
                c.section, c.section_title,
                c.page_start, c.page_end,
                c.mart_name,
                c.topics, c.relevance_tags,
                c.embedding,
            )
            for c in chunks
        ]
        execute_values(cur, f"""
            INSERT INTO {chunks_t}
                (chunk_id, doc_id, content, content_tokens, chunk_seq, chunk_type,
                 chapter, chapter_title, section, section_title,
                 page_start, page_end, mart_name, topics, relevance_tags, embedding)
            VALUES %s
        """, chunk_rows, template="""(
            %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s::text[], %s::text[], %s::vector
        )""")

        conn.commit()

    log.info(f"  ✓ Stored {len(chunks)} chunks for '{doc.title}'")
    return len(chunks)


# ═════════════════════════════════════════════════════════════════════════════
# DOCUMENT REGISTRY
# ═════════════════════════════════════════════════════════════════════════════

DOCUMENTS = {
    # ── Existing documents ──────────────────────────────────────────────────
    "weiss_book": {
        "record": DocumentRecord(
            doc_id           = "weiss-after-the-trade-3ed",
            title            = "After the Trade Is Made: Processing Securities Transactions (3rd Ed.)",
            doc_type         = "book",
            source_file      = "After_the_Trade_Is_Made__Revised_Third_Ed_-_Processing_--_David_M__Weiss_-sd3.pdf",
            author           = "David M. Weiss",
            publisher        = "Portfolio / Penguin",
            publication_year = 2006,
            domain           = "fund_operations",
        ),
        "ingest_fn": "pdf_book",
        "path":      "data/raw/After_the_Trade_Is_Made__Revised_Third_Ed_-_Processing_--_David_M__Weiss_-sd3.pdf",
    },
    "custody_api": {
        "record": DocumentRecord(
            doc_id           = "sst-custody-accounting-v10",
            title            = "State Street Core Data Consumption Model: Custody Accounting v10",
            doc_type         = "api_spec",
            source_file      = "v10_State_Street_Core_Data_Consumption_Model__Custody_Accounting_2026_02_14.xlsx",
            author           = "State Street Corporation",
            publication_year = 2026,
            domain           = "fund_operations",
            metadata         = {"version": "10", "api_type": "custody_accounting"},
        ),
        "ingest_fn": "excel",
        "path":      "data/raw/v10_State_Street_Core_Data_Consumption_Model__Custody_Accounting_2026_02_14.xlsx",
    },
    "performance_api": {
        "record": DocumentRecord(
            doc_id           = "sst-performance-api-v3",
            title            = "State Street Performance API v3",
            doc_type         = "api_spec",
            source_file      = "Performance_API_Updated_-_V3.xlsx",
            author           = "State Street Corporation",
            publication_year = 2026,
            domain           = "fund_operations",
            metadata         = {"version": "3", "api_type": "performance"},
        ),
        "ingest_fn": "excel",
        "path":      "data/raw/Performance_API_Updated_-_V3.xlsx",
    },

    # ── Template entries for future documents ───────────────────────────────
    # Copy, rename the key, fill in the fields, drop the file in data/raw/,
    # and run: python ingest.py --doc your_key

    # "dtcc_settlement_guide": {
    #     "record": DocumentRecord(
    #         doc_id           = "dtcc-settlement-guide-2024",
    #         title            = "DTCC Settlement Service Guide",
    #         doc_type         = "manual",          # ← routes to ingest_pdf_manual
    #         source_file      = "DTCC_Settlement_Service_Guide_2024.pdf",
    #         author           = "DTCC",
    #         publication_year = 2024,
    #         domain           = "fund_operations",
    #         metadata         = {"issuer": "DTCC", "service": "settlement"},
    #     ),
    #     "ingest_fn": "pdf_manual",
    #     "path":      "data/raw/DTCC_Settlement_Service_Guide_2024.pdf",
    # },

    # "nscc_rules_procedures": {
    #     "record": DocumentRecord(
    #         doc_id           = "nscc-rules-procedures-2024",
    #         title            = "NSCC Rules and Procedures",
    #         doc_type         = "manual",
    #         source_file      = "NSCC_Rules_Procedures_2024.pdf",
    #         author           = "NSCC / DTCC",
    #         publication_year = 2024,
    #         domain           = "fund_operations",
    #         metadata         = {"issuer": "NSCC", "type": "rules"},
    #     ),
    #     "ingest_fn": "pdf_manual",
    #     "path":      "data/raw/NSCC_Rules_Procedures_2024.pdf",
    # },

    # "paper_nav_errors": {
    #     "record": DocumentRecord(
    #         doc_id           = "paper-nav-errors-2023",
    #         title            = "Sources of NAV Error in Mutual Funds",
    #         doc_type         = "paper",            # ← routes to ingest_pdf_paper
    #         source_file      = "nav_errors_paper_2023.pdf",
    #         author           = "Author et al.",
    #         publication_year = 2023,
    #         domain           = "fund_operations",
    #         metadata         = {"journal": "Journal of Portfolio Management"},
    #     ),
    #     "ingest_fn": "pdf_paper",
    #     "path":      "data/raw/nav_errors_paper_2023.pdf",
    # },
}

# Dispatch table: ingest_fn string → function
INGEST_DISPATCH = {
    "pdf_book":   ingest_pdf_book,
    "pdf_manual": ingest_pdf_manual,
    "pdf_paper":  ingest_pdf_paper,
    "excel":      ingest_excel_api_spec,
}


# ═════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

def main():
    global DB_DSN, OPENAI_API_KEY
    DB_DSN, OPENAI_API_KEY = _get_config()

    parser = argparse.ArgumentParser(description="Otomeshon RAG ingestion pipeline (NAV RAG)")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--doc", choices=list(DOCUMENTS.keys()), help="Ingest a specific document")
    group.add_argument("--all", action="store_true", help="Ingest all documents")
    args = parser.parse_args()

    if not DB_DSN:
        raise EnvironmentError("Set OTOMESHON_DB_DSN or DATABASE_URL (or app config)")
    if not OPENAI_API_KEY:
        raise EnvironmentError("OPENAI_API_KEY environment variable or app config not set")

    oai_client = openai.OpenAI(api_key=OPENAI_API_KEY)
    conn       = psycopg2.connect(DB_DSN)

    docs_to_process = list(DOCUMENTS.keys()) if args.all else [args.doc]

    total = 0
    for doc_key in docs_to_process:
        entry    = DOCUMENTS[doc_key]
        doc      = entry["record"]
        path     = entry["path"]
        fn_key   = entry["ingest_fn"]

        log.info(f"\n{'='*60}")
        log.info(f"Ingesting: {doc.title}")
        log.info(f"Strategy:  {fn_key}")
        log.info(f"{'='*60}")

        if not os.path.exists(path):
            log.error(f"File not found: {path}")
            log.error("Place source files in data/raw/ before running ingestion.")
            continue

        ingest_fn = INGEST_DISPATCH.get(fn_key)
        if not ingest_fn:
            log.error(f"Unknown ingest_fn '{fn_key}' for document '{doc_key}'")
            continue

        n = ingest_fn(path, doc, oai_client, conn)
        total += n

    conn.close()
    log.info(f"\n✓ Ingestion complete. Total chunks stored: {total}")


if __name__ == "__main__":
    main()

